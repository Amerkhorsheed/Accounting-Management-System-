"""
Export Service - Report Export Functionality

This module provides export functionality for reports including:
- Excel export using openpyxl
- PDF export using reportlab with RTL support
- Print functionality using PySide6 print dialog

Requirements: 3.1, 3.2, 3.3, 3.4, 3.5, 3.6, 4.1, 4.2, 4.3, 4.4, 4.5, 4.6, 4.7, 5.1, 5.2, 5.3, 5.4, 5.5
"""
import os
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

from PySide6.QtWidgets import QFileDialog, QWidget
from PySide6.QtPrintSupport import QPrinter, QPrintPreviewDialog
from PySide6.QtGui import QTextDocument
from PySide6.QtCore import Qt

from ..config import config

# Configure logger
logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Exception raised for export errors."""
    
    def __init__(self, message: str, error_code: str = 'EXPORT_ERROR'):
        self.message = message
        self.error_code = error_code
        super().__init__(self.message)


class ExportService:
    """
    Service for exporting reports to Excel and PDF formats.
    
    Requirements:
    - 3.1, 3.2, 3.3, 3.4, 3.5, 3.6: Excel export functionality
    - 4.1, 4.2, 4.3, 4.4, 4.5, 4.6, 4.7: PDF export functionality
    - 5.1, 5.2, 5.3, 5.4, 5.5: Print functionality
    """
    
    @staticmethod
    def export_to_excel(
        data: List[Dict],
        columns: List[Tuple[str, str]],
        filename: str,
        title: str,
        parent: QWidget = None,
        summary: Dict[str, Any] = None
    ) -> bool:
        """
        Export data to Excel file.
        
        Args:
            data: List of dictionaries containing row data
            columns: List of tuples (key, header_label) for columns
            filename: Default filename for save dialog
            title: Report title to include in the file
            parent: Parent widget for file dialog
            summary: Optional summary data to include at top
            
        Returns:
            True if export successful, False otherwise
            
        Requirements: 3.1, 3.2, 3.3, 3.4, 3.5, 3.6
        """
        try:
            # Import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                from openpyxl.utils import get_column_letter
            except ImportError:
                raise ExportError(
                    "مكتبة Excel غير متوفرة. يرجى تثبيت openpyxl",
                    'LIBRARY_NOT_FOUND'
                )
            
            # Check if there's data to export
            if not data:
                raise ExportError("لا توجد بيانات للتصدير", 'NO_DATA')
            
            # Show save dialog
            # Requirements: 3.5 - Prompt user to save file
            file_path, _ = QFileDialog.getSaveFileName(
                parent,
                "حفظ ملف Excel",
                filename,
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return False  # User cancelled
            
            # Ensure .xlsx extension
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel sheet name limit
            
            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            header_font_white = Font(bold=True, size=12, color='FFFFFF')
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            rtl_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            current_row = 1
            
            # Add title
            # Requirements: 3.4 - Include report title
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            title_cell = ws.cell(row=current_row, column=1, value=title)
            title_cell.font = title_font
            title_cell.alignment = center_alignment
            current_row += 1
            
            # Add generation date
            # Requirements: 3.4 - Include generation date
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            date_cell = ws.cell(
                row=current_row, 
                column=1, 
                value=f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
            date_cell.alignment = center_alignment
            current_row += 2
            
            # Add summary if provided
            if summary:
                for key, value in summary.items():
                    ws.cell(row=current_row, column=1, value=key)
                    ws.cell(row=current_row, column=2, value=str(value))
                    current_row += 1
                current_row += 1
            
            # Add headers
            # Requirements: 3.2 - Include all visible data columns
            for col_idx, (key, header) in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Add data rows
            for row_data in data:
                for col_idx, (key, _) in enumerate(columns, 1):
                    value = row_data.get(key, '')
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = rtl_alignment
                    
                    # Format numbers with proper decimal places
                    # Requirements: 3.3 - Format numbers properly
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                
                current_row += 1
            
            # Auto-adjust column widths
            for col_idx, (key, header) in enumerate(columns, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(header))
                for row in ws.iter_rows(min_row=current_row - len(data), max_row=current_row - 1, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Set RTL direction for the sheet
            ws.sheet_view.rightToLeft = True
            
            # Save workbook
            wb.save(file_path)
            logger.info(f"Excel export successful: {file_path}")
            return True
            
        except ExportError:
            raise
        except PermissionError:
            # Requirements: 3.6 - Handle file permission errors
            raise ExportError(
                "لا يمكن حفظ الملف. تحقق من الصلاحيات أو أغلق الملف إذا كان مفتوحاً",
                'PERMISSION_DENIED'
            )
        except Exception as e:
            logger.exception(f"Excel export failed: {e}")
            raise ExportError(f"فشل تصدير Excel: {str(e)}", 'EXPORT_FAILED')


    @staticmethod
    def export_to_pdf(
        data: List[Dict],
        columns: List[Tuple[str, str]],
        filename: str,
        title: str,
        parent: QWidget = None,
        company_info: Dict[str, str] = None,
        summary: Dict[str, Any] = None,
        date_range: Tuple[str, str] = None
    ) -> bool:
        """
        Export data to PDF file with RTL support.
        
        Args:
            data: List of dictionaries containing row data
            columns: List of tuples (key, header_label) for columns
            filename: Default filename for save dialog
            title: Report title
            parent: Parent widget for file dialog
            company_info: Optional company header information
            summary: Optional summary data
            date_range: Optional tuple of (start_date, end_date)
            
        Returns:
            True if export successful, False otherwise
            
        Requirements: 4.1, 4.2, 4.3, 4.4, 4.5, 4.6, 4.7
        """
        try:
            # Import reportlab
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import cm, mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            except ImportError:
                raise ExportError(
                    "مكتبة PDF غير متوفرة. يرجى تثبيت reportlab",
                    'LIBRARY_NOT_FOUND'
                )
            
            # Check if there's data to export
            if not data:
                raise ExportError("لا توجد بيانات للتصدير", 'NO_DATA')
            
            # Show save dialog
            # Requirements: 4.6 - Prompt user to save file
            file_path, _ = QFileDialog.getSaveFileName(
                parent,
                "حفظ ملف PDF",
                filename,
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return False  # User cancelled
            
            # Ensure .pdf extension
            if not file_path.endswith('.pdf'):
                file_path += '.pdf'
            
            # Try to register Arabic font
            try:
                # Try common Arabic fonts
                arabic_fonts = [
                    ('C:/Windows/Fonts/arial.ttf', 'Arial'),
                    ('C:/Windows/Fonts/tahoma.ttf', 'Tahoma'),
                    ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 'DejaVuSans'),
                ]
                font_name = 'Helvetica'  # Default fallback
                
                for font_path, name in arabic_fonts:
                    if os.path.exists(font_path):
                        try:
                            pdfmetrics.registerFont(TTFont(name, font_path))
                            font_name = name
                            break
                        except:
                            continue
            except Exception as e:
                logger.warning(f"Could not register Arabic font: {e}")
                font_name = 'Helvetica'
            
            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Create custom styles for RTL
            # Requirements: 4.3 - Format with proper Arabic text direction (RTL)
            title_style = ParagraphStyle(
                'TitleRTL',
                parent=styles['Title'],
                fontName=font_name,
                fontSize=16,
                alignment=TA_CENTER,
                spaceAfter=12
            )
            
            header_style = ParagraphStyle(
                'HeaderRTL',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=12,
                alignment=TA_CENTER,
                spaceAfter=6
            )
            
            normal_style = ParagraphStyle(
                'NormalRTL',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                alignment=TA_RIGHT
            )
            
            # Add company header
            # Requirements: 4.2 - Include company header information
            if company_info is None:
                company_info = {
                    'name': config.COMPANY_NAME,
                    'address': config.COMPANY_ADDRESS,
                    'phone': config.COMPANY_PHONE
                }
            
            if company_info.get('name'):
                elements.append(Paragraph(company_info['name'], title_style))
            if company_info.get('address'):
                elements.append(Paragraph(company_info['address'], header_style))
            if company_info.get('phone'):
                elements.append(Paragraph(f"هاتف: {company_info['phone']}", header_style))
            
            elements.append(Spacer(1, 0.5*cm))
            
            # Add report title
            # Requirements: 4.4 - Include report title
            elements.append(Paragraph(title, title_style))
            
            # Add date range if provided
            # Requirements: 4.4 - Include date range
            if date_range:
                date_text = f"الفترة: من {date_range[0]} إلى {date_range[1]}"
                elements.append(Paragraph(date_text, header_style))
            
            # Add generation timestamp
            # Requirements: 4.4 - Include generation timestamp
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
            elements.append(Paragraph(f"تاريخ الإنشاء: {timestamp}", header_style))
            
            elements.append(Spacer(1, 0.5*cm))
            
            # Add summary if provided
            if summary:
                summary_data = [[str(k), str(v)] for k, v in summary.items()]
                summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
                summary_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 0.5*cm))
            
            # Prepare table data
            # Requirements: 4.5 - Format tables with proper borders and alignment
            # Reverse columns for RTL display
            reversed_columns = list(reversed(columns))
            table_data = [[header for _, header in reversed_columns]]
            
            for row_data in data:
                row = []
                for key, _ in reversed_columns:
                    value = row_data.get(key, '')
                    # Format numbers
                    if isinstance(value, (int, float)):
                        value = f"{value:,.2f}"
                    row.append(str(value) if value is not None else '')
                table_data.append(row)
            
            # Calculate column widths
            available_width = A4[0] - 3*cm
            col_width = available_width / len(columns)
            col_widths = [col_width] * len(columns)
            
            # Create table
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Apply table style
            table_style = TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # Borders
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                # Padding
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ])
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            logger.info(f"PDF export successful: {file_path}")
            return True
            
        except ExportError:
            raise
        except PermissionError:
            # Requirements: 4.7 - Handle file permission errors
            raise ExportError(
                "لا يمكن حفظ الملف. تحقق من الصلاحيات أو أغلق الملف إذا كان مفتوحاً",
                'PERMISSION_DENIED'
            )
        except Exception as e:
            logger.exception(f"PDF export failed: {e}")
            raise ExportError(f"فشل إنشاء ملف PDF: {str(e)}", 'EXPORT_FAILED')


    @staticmethod
    def print_document(
        html_content: str,
        parent: QWidget = None,
        title: str = "طباعة"
    ) -> bool:
        """
        Print HTML document using print preview dialog.
        
        Args:
            html_content: HTML content to print
            parent: Parent widget for print dialog
            title: Document title
            
        Returns:
            True if print was initiated, False if cancelled
            
        Requirements: 5.1, 5.2, 5.3, 5.4, 5.5
        """
        try:
            # Create printer
            printer = QPrinter(QPrinter.HighResolution)
            printer.setDocName(title)
            
            # Create text document with HTML content
            # Requirements: 5.2 - Display in print-friendly format
            document = QTextDocument()
            document.setHtml(html_content)
            
            # Show print preview dialog
            # Requirements: 5.1 - Open print preview dialog
            preview = QPrintPreviewDialog(printer, parent)
            preview.setWindowTitle(f"معاينة الطباعة - {title}")
            
            # Connect paint request to document print
            preview.paintRequested.connect(lambda p: document.print_(p))
            
            # Show dialog and check result
            result = preview.exec()
            
            if result == QPrintPreviewDialog.Accepted:
                logger.info(f"Print initiated: {title}")
                return True
            
            return False
            
        except Exception as e:
            logger.exception(f"Print failed: {e}")
            raise ExportError(f"فشل الطباعة: {str(e)}", 'PRINT_FAILED')

    @staticmethod
    def generate_statement_html(
        customer_info: Dict[str, Any],
        statement_data: Dict[str, Any],
        transactions: List[Dict],
        company_info: Dict[str, str] = None,
        date_range: Tuple[str, str] = None
    ) -> str:
        """
        Generate print-friendly HTML for customer statement.
        
        Args:
            customer_info: Customer details (name, code, etc.)
            statement_data: Statement summary (opening_balance, closing_balance, etc.)
            transactions: List of transactions
            company_info: Optional company header information
            date_range: Optional tuple of (start_date, end_date)
            
        Returns:
            HTML string for printing
            
        Requirements: 5.2, 5.3, 5.4
        """
        if company_info is None:
            company_info = {
                'name': config.COMPANY_NAME,
                'address': config.COMPANY_ADDRESS,
                'phone': config.COMPANY_PHONE
            }
        
        # Build HTML
        html = f"""
        <!DOCTYPE html>
        <html dir="rtl" lang="ar">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
                    direction: rtl;
                    padding: 20px;
                    font-size: 12px;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 20px;
                    border-bottom: 2px solid #2563EB;
                    padding-bottom: 10px;
                }}
                .company-name {{
                    font-size: 18px;
                    font-weight: bold;
                    color: #2563EB;
                }}
                .company-info {{
                    font-size: 11px;
                    color: #666;
                }}
                .title {{
                    font-size: 16px;
                    font-weight: bold;
                    text-align: center;
                    margin: 15px 0;
                }}
                .customer-info {{
                    background: #f8f9fa;
                    padding: 10px;
                    border-radius: 5px;
                    margin-bottom: 15px;
                }}
                .customer-info table {{
                    width: 100%;
                }}
                .customer-info td {{
                    padding: 3px 10px;
                }}
                .summary {{
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 15px;
                    background: #e3f2fd;
                    padding: 10px;
                    border-radius: 5px;
                }}
                .summary-item {{
                    text-align: center;
                }}
                .summary-label {{
                    font-size: 10px;
                    color: #666;
                }}
                .summary-value {{
                    font-size: 14px;
                    font-weight: bold;
                }}
                .debit {{ color: #EF4444; }}
                .credit {{ color: #10B981; }}
                table.transactions {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                }}
                table.transactions th {{
                    background: #2563EB;
                    color: white;
                    padding: 8px;
                    text-align: center;
                    font-size: 11px;
                }}
                table.transactions td {{
                    border: 1px solid #ddd;
                    padding: 6px;
                    text-align: center;
                    font-size: 10px;
                }}
                table.transactions tr:nth-child(even) {{
                    background: #f8f9fa;
                }}
                .footer {{
                    margin-top: 20px;
                    text-align: center;
                    font-size: 10px;
                    color: #666;
                    border-top: 1px solid #ddd;
                    padding-top: 10px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="company-name">{company_info.get('name', '')}</div>
                <div class="company-info">{company_info.get('address', '')}</div>
                <div class="company-info">هاتف: {company_info.get('phone', '')}</div>
            </div>
            
            <div class="title">كشف حساب العميل</div>
        """
        
        # Add date range if provided
        if date_range:
            html += f'<div style="text-align: center; margin-bottom: 10px;">الفترة: من {date_range[0]} إلى {date_range[1]}</div>'
        
        # Customer info section
        # Requirements: 5.3 - Include customer details
        html += f"""
            <div class="customer-info">
                <table>
                    <tr>
                        <td><strong>اسم العميل:</strong></td>
                        <td>{customer_info.get('name', '-')}</td>
                        <td><strong>كود العميل:</strong></td>
                        <td>{customer_info.get('code', '-')}</td>
                    </tr>
                </table>
            </div>
        """
        
        # Summary section
        # Requirements: 5.4 - Include opening balance, closing balance, and totals
        opening = float(statement_data.get('opening_balance', 0))
        closing = float(statement_data.get('closing_balance', 0))
        total_debit = float(statement_data.get('total_invoices', 0))
        total_payments = float(statement_data.get('total_payments', 0))
        total_returns = float(statement_data.get('total_returns', 0))
        total_credit = total_payments + total_returns
        
        html += f"""
            <table style="width: 100%; margin-bottom: 15px;">
                <tr>
                    <td style="text-align: center; padding: 10px; background: #f0f9ff; border-radius: 5px;">
                        <div class="summary-label">الرصيد الافتتاحي</div>
                        <div class="summary-value">{opening:,.2f} ل.س</div>
                    </td>
                    <td style="text-align: center; padding: 10px; background: #fef2f2; border-radius: 5px;">
                        <div class="summary-label">إجمالي المدين</div>
                        <div class="summary-value debit">{total_debit:,.2f} ل.س</div>
                    </td>
                    <td style="text-align: center; padding: 10px; background: #f0fdf4; border-radius: 5px;">
                        <div class="summary-label">إجمالي الدائن</div>
                        <div class="summary-value credit">{total_credit:,.2f} ل.س</div>
                    </td>
                    <td style="text-align: center; padding: 10px; background: #eff6ff; border-radius: 5px;">
                        <div class="summary-label">الرصيد الختامي</div>
                        <div class="summary-value" style="color: #2563EB;">{closing:,.2f} ل.س</div>
                    </td>
                </tr>
            </table>
        """
        
        # Transactions table
        # Requirements: 5.3 - Include transaction table
        html += """
            <table class="transactions">
                <thead>
                    <tr>
                        <th>التاريخ</th>
                        <th>النوع</th>
                        <th>المرجع</th>
                        <th>البيان</th>
                        <th>مدين</th>
                        <th>دائن</th>
                        <th>الرصيد</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        type_map = {
            'invoice': 'فاتورة',
            'payment': 'دفعة',
            'return': 'مرتجع',
            'opening': 'رصيد افتتاحي',
            'adjustment': 'تسوية'
        }
        
        for trans in transactions:
            trans_type = trans.get('type', '')
            type_display = type_map.get(trans_type, trans_type)
            debit = float(trans.get('debit', 0))
            credit = float(trans.get('credit', 0))
            balance = float(trans.get('balance', 0))
            
            debit_str = f"{debit:,.2f}" if debit > 0 else "-"
            credit_str = f"{credit:,.2f}" if credit > 0 else "-"
            
            debit_class = 'debit' if debit > 0 else ''
            credit_class = 'credit' if credit > 0 else ''
            
            html += f"""
                <tr>
                    <td>{trans.get('date', '')}</td>
                    <td>{type_display}</td>
                    <td>{trans.get('reference', '')}</td>
                    <td>{trans.get('description', '')}</td>
                    <td class="{debit_class}">{debit_str}</td>
                    <td class="{credit_class}">{credit_str}</td>
                    <td><strong>{balance:,.2f}</strong></td>
                </tr>
            """
        
        html += """
                </tbody>
            </table>
        """
        
        # Footer
        from datetime import datetime
        html += f"""
            <div class="footer">
                تم إنشاء هذا التقرير بتاريخ {datetime.now().strftime('%Y-%m-%d %H:%M')}
            </div>
        </body>
        </html>
        """
        
        return html

    @staticmethod
    def generate_report_html(
        title: str,
        data: List[Dict],
        columns: List[Tuple[str, str]],
        company_info: Dict[str, str] = None,
        summary: Dict[str, Any] = None,
        date_range: Tuple[str, str] = None
    ) -> str:
        """
        Generate print-friendly HTML for a generic report.
        
        Args:
            title: Report title
            data: List of dictionaries containing row data
            columns: List of tuples (key, header_label) for columns
            company_info: Optional company header information
            summary: Optional summary data
            date_range: Optional tuple of (start_date, end_date)
            
        Returns:
            HTML string for printing
        """
        if company_info is None:
            company_info = {
                'name': config.COMPANY_NAME,
                'address': config.COMPANY_ADDRESS,
                'phone': config.COMPANY_PHONE
            }
        
        # Build HTML
        html = f"""
        <!DOCTYPE html>
        <html dir="rtl" lang="ar">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
                    direction: rtl;
                    padding: 20px;
                    font-size: 12px;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 20px;
                    border-bottom: 2px solid #2563EB;
                    padding-bottom: 10px;
                }}
                .company-name {{
                    font-size: 18px;
                    font-weight: bold;
                    color: #2563EB;
                }}
                .company-info {{
                    font-size: 11px;
                    color: #666;
                }}
                .title {{
                    font-size: 16px;
                    font-weight: bold;
                    text-align: center;
                    margin: 15px 0;
                }}
                .summary {{
                    background: #f8f9fa;
                    padding: 10px;
                    border-radius: 5px;
                    margin-bottom: 15px;
                }}
                .summary table {{
                    width: 100%;
                }}
                .summary td {{
                    padding: 5px 10px;
                }}
                table.data {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                }}
                table.data th {{
                    background: #2563EB;
                    color: white;
                    padding: 8px;
                    text-align: center;
                    font-size: 11px;
                }}
                table.data td {{
                    border: 1px solid #ddd;
                    padding: 6px;
                    text-align: center;
                    font-size: 10px;
                }}
                table.data tr:nth-child(even) {{
                    background: #f8f9fa;
                }}
                .footer {{
                    margin-top: 20px;
                    text-align: center;
                    font-size: 10px;
                    color: #666;
                    border-top: 1px solid #ddd;
                    padding-top: 10px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="company-name">{company_info.get('name', '')}</div>
                <div class="company-info">{company_info.get('address', '')}</div>
                <div class="company-info">هاتف: {company_info.get('phone', '')}</div>
            </div>
            
            <div class="title">{title}</div>
        """
        
        # Add date range if provided
        if date_range:
            html += f'<div style="text-align: center; margin-bottom: 10px;">الفترة: من {date_range[0]} إلى {date_range[1]}</div>'
        
        # Summary section
        if summary:
            html += '<div class="summary"><table>'
            for key, value in summary.items():
                html += f'<tr><td><strong>{key}:</strong></td><td>{value}</td></tr>'
            html += '</table></div>'
        
        # Data table
        html += '<table class="data"><thead><tr>'
        for _, header in columns:
            html += f'<th>{header}</th>'
        html += '</tr></thead><tbody>'
        
        for row_data in data:
            html += '<tr>'
            for key, _ in columns:
                value = row_data.get(key, '')
                if isinstance(value, (int, float)):
                    value = f"{value:,.2f}"
                html += f'<td>{value}</td>'
            html += '</tr>'
        
        html += '</tbody></table>'
        
        # Footer
        from datetime import datetime
        html += f"""
            <div class="footer">
                تم إنشاء هذا التقرير بتاريخ {datetime.now().strftime('%Y-%m-%d %H:%M')}
            </div>
        </body>
        </html>
        """
        
        return html


# Create singleton instance
export_service = ExportService()
